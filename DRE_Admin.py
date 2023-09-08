import openpyxl as px
import os

# Directory containing the Excel files
folder_path = r'\\ban-c-000yr\C$\Users\DNH4KOR\OneDrive - Bosch Group\DRE'
# Create a new workbook to store the merged data
merged_wb = px.Workbook()
merged_sheet = merged_wb.active
merged_row = 1

# Iterate through all files in the folder
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        print(file_path)

        # Open the Excel file
        wb = px.load_workbook(file_path)
        sheet = wb.active

        # Copy data from the file, starting from row 10
        for row in sheet.iter_rows(min_row=10):
            merged_sheet.append([cell.value for cell in row])

            # Get the value from cell C3
            c3_value = sheet['C3'].value
            merged_row = merged_sheet.max_row
           # merged_row += 1
            # Print the C3 value in the respective A column row in the merged sheet
            merged_sheet.cell(row=merged_row, column=1, value=c3_value)

        # Close the file
        wb.close()

# Save the new merged Excel file
merged_file_path = 'C:/Users/DNH4KOR/Desktop/merged_output95.xlsx'  # Replace with the desired output file path
merged_wb.save(merged_file_path)

# Close the merged workbook
merged_wb.close()

# Open the merged Excel file
wb = px.load_workbook(merged_file_path)
merged_sheet = wb.active

# Iterate through all rows in the sheet in reverse order to safely delete rows
for row in reversed(list(merged_sheet.iter_rows(min_row=1, max_row=merged_sheet.max_row))):
    # Check if all cells except the first one are empty
    if all(cell.value is None for cell in row[1:]):
        merged_sheet.delete_rows(row[0].row)

# Save the modified Excel file
output_file_path = 'C:/Users/DNH4KOR/Desktop/output_excel_file_95.xlsx'  # Replace with the desired output file path
wb.save(output_file_path)
print("your output file is ready")
# Close the workbook
wb.close()